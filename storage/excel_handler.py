"""Excel export handler for qualified leads."""

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from models.lead import Lead


def export_to_excel(
    leads: list[Lead], 
    qualifications: list[dict], 
    filename: str
) -> None:
    """
    Export qualified leads to Excel with formatting.
    
    Args:
        leads: List of Lead objects
        qualifications: List of qualification dicts (must match leads order)
        filename: Output Excel file path
    """
    if len(leads) != len(qualifications):
        raise ValueError(f"Leads ({len(leads)}) and qualifications ({len(qualifications)}) must have same length")
    
    # Ensure directory exists
    Path(filename).parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Qualified Leads"
    
    # Define headers
    headers = [
        "Author",
        "Source", 
        "Content",
        "URL",
        "Engagement Score",
        "Is Qualified",
        "Confidence",
        "Reason",
        "Service Match",
        "Timestamp"
    ]
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Combine leads with qualifications and sort by confidence score descending
    combined = list(zip(leads, qualifications))
    combined.sort(key=lambda x: x[1].get('confidence_score', 0.0), reverse=True)
    
    # Define fills for conditional formatting
    qualified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    not_qualified_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    
    # Write data rows
    for row_idx, (lead, qual) in enumerate(combined, 2):
        # Truncate content to 200 chars
        content = lead.content[:200] + "..." if len(lead.content) > 200 else lead.content
        
        # Format service match as comma-separated string
        service_match = ", ".join(qual.get('service_match', []))
        
        # Format timestamp
        timestamp_str = lead.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        
        # Prepare row data
        row_data = [
            lead.author,
            lead.source,
            content,
            lead.url,
            lead.engagement_score,
            "Yes" if qual.get('is_qualified', False) else "No",
            round(qual.get('confidence_score', 0.0), 2),
            qual.get('reason', ''),
            service_match,
            timestamp_str
        ]
        
        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in [3, 8]))  # Wrap text for Content and Reason
            
            # Apply conditional formatting to entire row
            if qual.get('is_qualified', False):
                cell.fill = qualified_fill
            else:
                cell.fill = not_qualified_fill
    
    # Auto-adjust column widths
    column_widths = {
        1: 20,   # Author
        2: 12,   # Source
        3: 50,   # Content
        4: 40,   # URL
        5: 15,   # Engagement Score
        6: 12,   # Is Qualified
        7: 12,   # Confidence
        8: 50,   # Reason
        9: 30,   # Service Match
        10: 20   # Timestamp
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(filename)
    
    qualified_count = sum(1 for _, q in combined if q.get('is_qualified', False))
    print(f"\n✅ Exported {len(leads)} leads to {filename}")
    print(f"   • Qualified: {qualified_count}")
    print(f"   • Not Qualified: {len(leads) - qualified_count}")
    print(f"   • Sorted by confidence score (highest first)")


def export_qualified_only(
    leads: list[Lead],
    qualifications: list[dict],
    filename: str,
    min_confidence: float = 0.0
) -> None:
    """
    Export only qualified leads to Excel.
    
    Args:
        leads: List of Lead objects
        qualifications: List of qualification dicts (must match leads order)
        filename: Output Excel file path
        min_confidence: Minimum confidence score to include (default: 0.0)
    """
    # Filter to qualified leads only
    filtered = [
        (lead, qual) 
        for lead, qual in zip(leads, qualifications)
        if qual.get('is_qualified', False) and qual.get('confidence_score', 0.0) >= min_confidence
    ]
    
    if not filtered:
        print(f"⚠️  No qualified leads found with confidence >= {min_confidence}")
        return
    
    qualified_leads, qualified_quals = zip(*filtered)
    
    print(f"📊 Exporting {len(qualified_leads)} qualified leads (confidence >= {min_confidence})...")
    export_to_excel(list(qualified_leads), list(qualified_quals), filename)


def export_by_service(
    leads: list[Lead],
    qualifications: list[dict],
    service: str,
    filename: str,
    min_confidence: float = 0.0
) -> None:
    """
    Export leads matching specific service to Excel.
    
    Args:
        leads: List of Lead objects
        qualifications: List of qualification dicts (must match leads order)
        service: Service to filter by ('RWA', 'Crypto/Blockchain', 'AI/ML')
        filename: Output Excel file path
        min_confidence: Minimum confidence score to include (default: 0.0)
    """
    # Filter to specific service
    filtered = [
        (lead, qual)
        for lead, qual in zip(leads, qualifications)
        if qual.get('is_qualified', False)
        and service in qual.get('service_match', [])
        and qual.get('confidence_score', 0.0) >= min_confidence
    ]
    
    if not filtered:
        print(f"⚠️  No qualified leads found for service '{service}' with confidence >= {min_confidence}")
        return
    
    service_leads, service_quals = zip(*filtered)
    
    print(f"📊 Exporting {len(service_leads)} {service} leads (confidence >= {min_confidence})...")
    export_to_excel(list(service_leads), list(service_quals), filename)
